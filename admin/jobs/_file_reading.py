"""admin.jobs._file_reading — CSV/Excel parsing shared between
preview_import_columns (admin/api/data_import_api.py) and the actual
import job (admin/jobs/data_import_job.py), so the two can never sniff
the format or read cell values differently from one another.

Format is sniffed off the ORIGINAL filename's extension, not the
uploaded content-type — filer's own upload() only maps a handful of
content types to a real extension (csv isn't one of them, see its
_EXTENSION_FOR), and different browsers/OSes report different content
types for the same .csv file anyway. The filename extension is stable
and always available (filerfile.original_filename).

read_all_rows() streams: it hands back an async row ITERATOR, not a
materialized list — the earlier shape (`list[list[Any]]`) meant the
import job's own bookkeeping pass, which already batches its WRITES in
bounded chunks, was still reading the whole file into memory first
regardless. XLSX goes through a temp file on disk rather than an
in-memory BytesIO for the identical reason: openpyxl's read_only mode
already streams row data once it has a seekable source, but only if that
source isn't itself a giant in-memory buffer — a temp file also bounds
"how much RAM can a single malicious upload force us to hold" to roughly
one chunk at a time instead of the whole (possibly zip-bomb-expanded)
archive. Exactly ONE read/stream of the underlying file per call — the
header comes off the SAME pass as the data rows, not a separate peek."""

from __future__ import annotations

import csv
import io
import os
import tempfile
from pathlib import Path
from typing import Any, AsyncIterator

#: Matches filer.providers' own StorageProvider.read_stream default —
#: not imported (that's a Protocol, no constant to share), just the same
#: number for the same "reasonable I/O chunk" reasoning.
_STREAM_CHUNK_SIZE = 65536
_PREVIEW_SAMPLE_ROWS = 10


def _is_xlsx(file_row: dict) -> bool:
    return file_row["original_filename"].lower().endswith(".xlsx")


def _provider(file_row: dict):
    from filer.providers import PROVIDERS

    return PROVIDERS[file_row["storage"]]


async def _read_bytes(file_row: dict) -> bytes:
    return await _provider(file_row).read(file_row["storage_key"])


async def _stream_to_tempfile(file_row: dict) -> Path:
    """Copies the stored file to a local temp file via the provider's own
    read_stream() (bounded per-chunk memory the whole way, including over
    S3 — StorageProvider.read_stream is part of the same Protocol either
    backend implements), rather than one `provider.read()` call handing
    back the entire file as a single in-memory bytes object."""
    fd, raw_path = tempfile.mkstemp(prefix="arc-import-", suffix=".xlsx")
    path = Path(raw_path)
    with os.fdopen(fd, "wb") as f:
        async for chunk in _provider(file_row).read_stream(
            file_row["storage_key"], chunk_size=_STREAM_CHUNK_SIZE
        ):
            f.write(chunk)
    return path


async def iter_rows_preview(file_row: dict) -> tuple[list[str], list[list[str]], int | None]:
    """(columns, up to _PREVIEW_SAMPLE_ROWS sample rows as strings,
    row_count_hint). row_count_hint is a real count for CSV (cheap — it's
    already fully in memory to get the header) and None for xlsx (would
    need a second full pass to count with openpyxl's own API; not worth
    it just for a preview hint)."""
    if _is_xlsx(file_row):
        path = await _stream_to_tempfile(file_row)
        try:
            import openpyxl

            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = [str(c) if c is not None else "" for c in next(rows_iter)]
            except StopIteration:
                return [], [], 0
            samples = []
            for i, row in enumerate(rows_iter):
                if i >= _PREVIEW_SAMPLE_ROWS:
                    break
                samples.append(["" if v is None else str(v) for v in row])
            wb.close()
            return header, samples, None
        finally:
            path.unlink(missing_ok=True)

    content = await _read_bytes(file_row)
    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    try:
        header = next(reader)
    except StopIteration:
        return [], [], 0
    samples = []
    total = 0
    for row in reader:
        total += 1
        if len(samples) < _PREVIEW_SAMPLE_ROWS:
            samples.append(row)
    return header, samples, total


async def _iter_csv_rows(reader: csv.reader) -> AsyncIterator[list[str]]:
    """Thin async wrapper around an already-open csv.reader's remaining
    rows — csv.reader is already a lazy, one-row-at-a-time iterator; the
    old code's only actual mistake was collecting its output into
    `list(reader)`. No `await` inside the loop (row iteration is pure,
    fast CPU work over an already-decoded in-memory string, not I/O), so
    this never yields control mid-row — it's an async generator purely so
    callers can `async for` it alongside the xlsx path uniformly."""
    for row in reader:
        yield row


async def _iter_xlsx_rows(rows_iter, path: Path, wb) -> AsyncIterator[list[Any]]:
    """Same shape as _iter_csv_rows, for openpyxl's own lazy row
    iterator. Owns closing the workbook AND deleting the temp file once
    exhausted (or once the caller stops iterating early and this
    generator is garbage-collected/closed — the `finally` still runs via
    PEP 342 generator close semantics)."""
    try:
        for row in rows_iter:
            yield ["" if v is None else v for v in row]
    finally:
        wb.close()
        path.unlink(missing_ok=True)


async def read_all_rows(file_row: dict) -> tuple[list[str], AsyncIterator[list[Any]]]:
    """(columns, an async iterator over every remaining data row's raw
    cell values) — used once by the import job's own first-run
    bookkeeping pass (see data_import_job.py), which already writes in
    bounded _BOOKKEEPING_BATCH-sized chunks; this is what makes the READ
    side of that pass bounded-memory too, instead of materializing the
    whole file's rows first and batching only the writes.

    Exactly one read/stream of the file — the header is pulled off the
    SAME open reader/workbook the returned iterator continues from, not
    a separate peek that would double the I/O."""
    if _is_xlsx(file_row):
        path = await _stream_to_tempfile(file_row)
        import openpyxl

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = [str(c) if c is not None else "" for c in next(rows_iter)]
        except StopIteration:
            wb.close()
            path.unlink(missing_ok=True)

            async def _empty() -> AsyncIterator[list[Any]]:
                return
                yield  # pragma: no cover - makes this a generator function

            return [], _empty()
        return header, _iter_xlsx_rows(rows_iter, path, wb)

    content = await _read_bytes(file_row)
    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    try:
        header = next(reader)
    except StopIteration:
        async def _empty() -> AsyncIterator[list[Any]]:
            return
            yield  # pragma: no cover - makes this a generator function

        return [], _empty()
    return header, _iter_csv_rows(reader)
